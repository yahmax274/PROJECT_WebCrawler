def input(file):
    from openpyxl import load_workbook
    wb = load_workbook(file)
    # Work Sheet
    ws = wb.get_sheet_by_name('sheet1')
    # Column
    column = ws['A']
    link_list = [column[x].value for x in range(len(column))]
    return (link_list)
def main(link):
    all_list.append(link*link)
    return all_list
def output(all_list):    
    import pandas as pd
    df = pd.DataFrame(all_list)
    df.to_csv("Index.csv",encoding='utf-8-Sig',index=False)
if __name__ == '__main__':
    import time
    all_list = []
    Href_list=[]
    start = time.time() # 開始測量執行時間
    link_list = input("NumberIndex.xlsx")#匯入檔案名稱
    for link in link_list[0:]:
        Href_list.append(link)
        main(link)
    print(len(all_list))
    output(all_list)
    end = time.time() # 結束測量執行時間
    print("執行時間為 %f 秒" % (end - start))