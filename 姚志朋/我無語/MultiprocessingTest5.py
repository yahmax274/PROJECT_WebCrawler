#數量錯誤
import requests
from bs4 import BeautifulSoup
import time
import datetime
import pandas as pd
import numpy
import openpyxl
from openpyxl import load_workbook
import concurrent.futures
import threading
import multiprocessing as mp

#list
type_list = []
brand_list = []
name_list = []
price_list = []
date_list = []
spec_list = []
act_list = []
all_list = []
Href_list=[]
def MultThread():
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(Main, Href_list)
def MultProcess():
    # with concurrent.futures.ProcessPoolExecutor(max_workers=5) as executor:executor.map(Main,Href_list, chunksize=5)
    with concurrent.futures.ProcessPoolExecutor() as executor:
        executor.map(Main,Href_list, chunksize=1000)

def WriteData():
    #匯出csv檔
    for i in range(len(Href_list)):
        all = type_list[i]+"|"+brand_list[i]+"|"+name_list[i]+"|"+price_list[i]+"|"+date_list[i]+"|"+spec_list[i]+"|"+act_list[i]
    all_list.append(all)
    raw_data ={"app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act": all_list}
    df = pd.DataFrame(raw_data,columns=["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"])
    df.to_csv("momo1.csv",encoding='utf-8-Sig',index=False)
def output(all_list):
    import pandas as pd
    raw_data ={"app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act": all_list}
    df = pd.DataFrame(raw_data,columns=["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"])
    df.to_csv("momo3.csv",encoding='utf-8-Sig',index=False)
def ReadData():
    fn = '3Ckey.xlsx'
    wb = load_workbook( fn )
    # Work Sheet
    ws = wb.get_sheet_by_name( 'sheet2' )
    # Column
    column = ws[ 'A' ]
    mylist = [column[x].value for x in range(len(column))]
    return mylist
def Main(Href_list):
    # url = "https://www.momoshop.com.tw/goods/GoodsDetail.jsp?i_code=10604406&sourcePageType=4"
    # url = Href_list
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36"}
    re = requests.get(Href_list,headers=headers)
    soup = BeautifulSoup(re.text,"html.parser")
    #type
    types = soup.find_all("div",{"id":"bt_2_layout_NAV"})
    try:
        for type in types:
            type = type.text
            type = type.replace(" ","")
            type = type.split("\n")
            type = type[2:]
            if type == []:
                continue
            else:
                for i in range(4):
                    if type[i] == "":
                        type[i] ="null"
                type = type[0]+"|"+type[1]+"|"+type[2]+"|"+type[3]+"|"+type[4]
                type_list.append(type)

        #brand
        brand = soup.find("a",{"class":"webBrandLink"}).get_text()
        brand_list.append(brand)
        #item_name
        name = soup.find("span",{"id":"osmGoodsName"}).get_text()
        name_list.append(name)
        #prices
        prices = soup.find_all("ul",class_ = "prdPrice")
        if prices ==[]:
                price = "null|null|null"
                price_list.append(price)
        else:
            for price in prices:
                price = price.text
                price = price.replace(" ","")
                price = price.split("\n")
                price = [i.strip() for i in price if i.strip()!=""]
                price_1 =""
                price_2 =""
                price_3 =""
                for i in range(len(price)):
                    if price_1 =="null" or price_1 =="":
                        if "市售價" in price[i]:
                            price_1 = price[i]
                            price_1 = price_1[3:]
                            price_1 = price_1.replace("元","")
                            price_1 = price_1.replace(",","")
                        else:
                            price_1 = "null"
                    if price_2 =="null" or price_2 =="":
                        if "促銷價" in price[i]:
                            price_2 = price[i]
                            price_2 = price_2[3:]
                            price_2 = price_2.replace("元","")
                            price_2 = price_2.replace("賣貴通報","")
                            price_2 = price_2.replace("下單再折","")
                            price_2 = price_2.replace(",","")
                        else:
                            price_2 = "null"
                    if price_3 =="null" or price_3 =="":
                        if "折扣後價格" in price[i]:
                            price_3 = price[i]
                            price_3 = price_3[5:]
                            price_3 = price_3.replace("元賣貴通報","")
                            price_3 = price_3.replace(",","")
                        else:
                            price_3 = "null"
                price = price_1+"|"+price_2+"|"+price_3
                price_list.append(price)
        #date
        date = datetime.date.today()
        date = str(date)
        date = date.replace("-","")
        date_list.append(date)
        #spec
        spec = soup.find("div",class_ = "vendordetailview specification").get_text()
        spec = spec.replace("\n","")
        spec = spec.replace(" ","")
        spec = spec.strip("TOP")
        spec_list.append(spec)
        #act
        try:
            act = soup.find("ul",class_ = "ineventArea").text
            act = act.replace(" ","")
            act = act.replace("\n","")
            act = act.replace("\xa0","")
            act = act.split("(說明)")
            act = act[:2]
            if act[1] =="":
                act = act[0]+"|null"
            else:
                act = act[0] +"|"+ act[1]
        except:
            act = "null|null"
        act_list.append(act)
    except:
        pass
    for i in range(len(type_list)):
        all = type_list[i]+"|"+brand_list[i]+"|"+name_list[i]+"|"+price_list[i]+"|"+date_list[i]+"|"+spec_list[i]+"|"+act_list[i]
        all_list.append(all)

    return(all_list)
if __name__ == '__main__':
    start = time.time() # 開始測量執行時間
    mylist=ReadData()
    for link in mylist[0:10]:##如果網址內無商品，會跳錯誤訊息
        Href = link
        Href_list.append(Href)
        # 無平行處理
        # Main(Href)
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(Main, Href_list)
    # MultThread()
    # MultProcess()
    # WriteData()
    # print(all_list)
    output(all_list)
    end = time.time() # 結束測量執行時間
    print("執行時間為 %f 秒" % (end - start))
    # print(Href_list)

