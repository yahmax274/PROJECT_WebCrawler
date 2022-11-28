#
def input(file):
    from openpyxl import load_workbook
    wb = load_workbook(file)
    # Work Sheet
    ws = wb.get_sheet_by_name('sheet1')
    # Column
    column = ws['A']
    link_list = [column[x].value for x in range(len(column))]
    return (link_list)

def main(link):#n為設定爬取資料筆數
    import requests
    from bs4 import BeautifulSoup
    import datetime

    url = link
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36"}
    re = requests.get(url,headers=headers)
    soup = BeautifulSoup(re.text,"html.parser")
    #type
    types = soup.find_all("div",{"id":"bt_2_layout_NAV"})
    try:
        for type in types:
            type = type.text
            type = type.replace(" ","")
            type = type.split("\n")
            type = type[2:]
            if type == []:
                continue
            else:
                for i in range(4):
                    if type[i] == "":
                        type[i] ="null"
                type = type[0]+"|"+type[1]+"|"+type[2]+"|"+type[3]+"|"+type[4]
                type_list.append(type)

        #brand
        brand = soup.find("a",{"class":"webBrandLink"}).get_text()
        brand_list.append(brand)
        #item_name
        name = soup.find("span",{"id":"osmGoodsName"}).get_text()
        name_list.append(name)
        #prices
        prices = soup.find_all("ul",class_ = "prdPrice")
        if prices ==[]:
                price = "null|null|null"
                price_list.append(price)
        else:
            for price in prices:
                price = price.text
                price = price.replace(" ","")
                price = price.split("\n")
                price = [i.strip() for i in price if i.strip()!=""]
                price_1 =""
                price_2 =""
                price_3 =""
                for i in range(len(price)):
                    if price_1 =="null" or price_1 =="":
                        if "市售價" in price[i]:
                            price_1 = price[i]
                            price_1 = price_1[3:]
                            price_1 = price_1.replace("元","")
                            price_1 = price_1.replace(",","")
                        else:
                            price_1 = "null"
                    if price_2 =="null" or price_2 =="":
                        if "促銷價" in price[i]:
                            price_2 = price[i]
                            price_2 = price_2[3:]
                            price_2 = price_2.replace("元","")
                            price_2 = price_2.replace("賣貴通報","")
                            price_2 = price_2.replace("下單再折","")
                            price_2 = price_2.replace(",","")
                        else:
                            price_2 = "null"
                    if price_3 =="null" or price_3 =="":
                        if "折扣後價格" in price[i]:
                            price_3 = price[i]
                            price_3 = price_3[5:]
                            price_3 = price_3.replace("元賣貴通報","")
                            price_3 = price_3.replace(",","")
                        else:
                            price_3 = "null"
                price = price_1+"|"+price_2+"|"+price_3
                price_list.append(price)
        #date
        date = datetime.date.today()
        date = str(date)
        date = date.replace("-","")
        date_list.append(date)
        #spec
        spec = soup.find("div",class_ = "vendordetailview specification").get_text()
        spec = spec.replace("\n","")
        spec = spec.replace(" ","")
        spec = spec.strip("TOP")
        spec_list.append(spec)
        #act
        try:
            act = soup.find("ul",class_ = "ineventArea").text
            act = act.replace(" ","")
            act = act.replace("\n","")
            act = act.replace("\xa0","")
            act = act.split("(說明)")
            act = act[:2]
            if act[1] =="":
                act = act[0]+"|null"
            else:
                act = act[0] +"|"+ act[1]
        except:
            act = "null|null"
        act_list.append(act)
    except:
        pass
    for i in range(len(type_list)):
        all = type_list[i]+"|"+brand_list[i]+"|"+name_list[i]+"|"+price_list[i]+"|"+date_list[i]+"|"+spec_list[i]+"|"+act_list[i]
    return(all)

#匯出csv檔
def output(all):
    import pandas as pd
    raw_data ={"app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act": all_list}
    df = pd.DataFrame(raw_data,columns=["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"])
    df.to_csv("momo3.csv",encoding='utf-8-Sig',index=False)

if __name__ == '__main__':
    import time
    import concurrent.futures
    #list
    type_list = []
    brand_list = []
    name_list = []
    price_list = []
    date_list = []
    spec_list = []
    act_list = []
    all_list = []
    start = time.time() # 開始測量執行時間
    link_list = input('3Ckey.xlsx')#匯入檔案名稱
    for link in link_list[:10]:#10為執行筆數
        try:
            all = main(link)
            all_list.append(all)
            # with concurrent.futures.ThreadPoolExecutor() as executor:
            #     all = executor.map(main, link)
            #     all_list.append(all)
        except:
            continue
    # for link in href_list:
    #     with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
    #         executor.map(main, link)
            # main(link)
    print(len(all_list))
    output(all_list)
    end = time.time() # 結束測量執行時間
    print("執行時間為 %f 秒" % (end - start))
