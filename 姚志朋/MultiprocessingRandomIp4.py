class IPCollector:#收集虛擬IP
    def __init__(self):
        self.usable_ips = []
        self.filtered_ips = []
    def collect_and_check_ip(self):#收集IP步驟1
        from fake_useragent import UserAgent
        import concurrent.futures
        import requests
        import re
        proxy_ips=[]
        response = requests.get("https://www.sslproxies.org/")#ip多但不穩定
        # response = requests.get("https://www.us-proxy.org/")#ip少但穩定
        proxy_ips = re.findall('\d+\.\d+\.\d+\.\d+:\d+', response.text) 

        with concurrent.futures.ProcessPoolExecutor() as executor:
            futures = [executor.submit(self.check_ip, ip) for ip in proxy_ips]
            for future in concurrent.futures.as_completed(futures):
                try:
                    usable_ip = future.result()
                    if usable_ip:
                        self.usable_ips.append(usable_ip)   
                except:
                    continue 
        print("收集到的IP數量:",len(self.usable_ips))

        user_agent = UserAgent()
        headers={ 'user-agent': user_agent.random }
        with concurrent.futures.ProcessPoolExecutor() as executor:#收集IP步驟3
            futures = [executor.submit(self.check_ip_momo, ip,headers) for ip in self.usable_ips]
            for future in concurrent.futures.as_completed(futures):
                try:
                    usable_ip = future.result()
                    if  usable_ip:
                        self.filtered_ips.append(usable_ip)    
                except:
                    continue 
        print("確定能使用的IP數量:",len(self.filtered_ips))
        print(self.filtered_ips)
        return self.filtered_ips

    def check_ip(self,ip):#收集IP步驟2
        import requests
        try:
            result = requests.get('https://ip.seeip.org/jsonip?',
                    proxies={'http': ip, 'https': ip},
                    timeout=5)
            if result.status_code == 200:
                return ip
        except:
            pass
        return None

    def check_ip_momo(self,ip,headers):#收集IP步驟4
        import requests
        try:
            result = requests.get('https://www.momoshop.com.tw/main/Main.jsp',
                    proxies={'http': ip, 'https': ip},
                    timeout=5,
                    headers=headers)
            if result.status_code==200:
                return ip
        except:
            pass
        return None
def Input(file):#匯入Excel資料
    from openpyxl import load_workbook
    wb = load_workbook(file)
    # Work Sheet
    ws = wb.get_sheet_by_name('sheet1')
    # Column
    column = ws['A']
    link_list = [column[x].value for x in range(len(column))]
    return (link_list) 
def DataCollect(link,proxy_ip):#主函式，抓取資料
    import requests
    from bs4 import BeautifulSoup
    import datetime
    from fake_useragent import UserAgent
    import random
    import time
    import pandas as pd
    delay_choices = [1,2,3,4,5]  #延遲的秒數
    delay = random.choice(delay_choices)
    time.sleep(delay)
    url = link
    ip=proxy_ip
    user_agent = UserAgent()
    headers={ 'user-agent': user_agent.random }
    try:
        re = requests.get(url,headers=headers, timeout=10,proxies={'http': f'{ip}', 'https': f'{ip}'})
        soup = BeautifulSoup(re.text,"html.parser")
            #type
        types = soup.find_all("div",{"id":"bt_2_layout_NAV"})
        if types != []:
            for type in types:
                type = type.text
                type = type.replace(" ","")
                type = type.split("\n")
                type = type[2:] 
                for i in range(4):  
                    if type[i] == "":
                        type[i] ="null"
                type = type[0]+"|"+type[1]+"|"+type[2]+"|"+type[3]+"|"+type[4] 
                type_list=type  
        try:     
            #brand
            brand = soup.find("a",{"class":"webBrandLink"}).get_text()
            brand_list=brand
            #item_name
            name = soup.find("span",{"id":"osmGoodsName"}).get_text()
            name_list=name
            #prices
            prices = soup.find_all("ul",class_ = "prdPrice")
            if prices ==[]:
                    price = "null|null|null"
                    price_list=price
            else:
                for price in prices:
                    price = price.text
                    price = price.replace(" ","")
                    price = price.split("\n")
                    price = [i.strip() for i in price if i.strip()!=""]
                    price_1 =""
                    price_2 =""
                    price_3 =""
                    for i in range(len(price)):
                        if price_1 =="null" or price_1 =="":
                            if "市售價" in price[i]:
                                price_1 = price[i]
                                price_1 = price_1[3:]
                                price_1 = price_1.replace("元","")   
                                price_1 = price_1.replace(",","")
                            else:
                                price_1 = "null"
                        if price_2 =="null" or price_2 =="":
                            if "促銷價" in price[i]:
                                price_2 = price[i]
                                price_2 = price_2[3:]
                                price_2 = price_2.replace("元","")
                                price_2 = price_2.replace("賣貴通報","")
                                price_2 = price_2.replace("下單再折","")
                                price_2 = price_2.replace(",","")
                            else:
                                price_2 = "null"  
                        if price_3 =="null" or price_3 =="":
                            if "折扣後價格" in price[i]:
                                price_3 = price[i]
                                price_3 = price_3[5:]
                                price_3 = price_3.replace("元賣貴通報","")
                                price_3 = price_3.replace(",","") 
                            else:
                                price_3 = "null"  
                    price = price_1+"|"+price_2+"|"+price_3
                    price_list=price
            #date
            date = datetime.date.today()
            date = str(date)
            date = date.replace("-","")
            date_list=date
            #spec
            spec = soup.find("div",class_ = "vendordetailview specification").get_text()
            spec = spec.replace("\n","")
            spec = spec.replace(" ","")
            spec = spec.strip("TOP")
            spec_list=spec
            #act
            try:
                act = soup.find("ul",class_ = "ineventArea").text
                act = act.replace(" ","")
                act = act.replace("\n","")
                act = act.replace("\xa0","")
                act = act.split("(說明)")
                act = act[:2]
                if act[1] =="":
                    act = act[0]+"|null"
                else:
                    act = act[0] +"|"+ act[1]       
            except:
                act = "null|null"
            act_list=act
            index = type_list+"|"+brand_list+"|"+name_list+"|"+price_list+"|"+date_list+"|"+spec_list+"|"+act_list
        except:
            index="null"
            pass
    except:
        raw_data = [url]
        df = pd.DataFrame(raw_data, columns=[""])
        df.to_csv("TimeoutURL.csv", mode='a', index=False, header=False,encoding='utf-8-Sig')
        pass
    raw_data = [index]
    df = pd.DataFrame(raw_data, columns=["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"])
    df.to_csv(save_file_name, mode='a', index=False, header=False,encoding='utf-8-Sig')
    return None
def MultiProcess1(link,proxy_ip):#平行處理主函式
    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(DataCollect,link,proxy_ip,chunksize=50) 
        executor.shutdown(wait=False)
def CcIndex():
    Index=len(Href_list)+Set_Number
    return Index
def Create_and_delete_file(save_file_name):#建立和刪除舊有檔案
    if os.path.exists(save_file_name):#刪除舊資料
        os.remove(save_file_name)
        raw_data = ["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"]
        df = pd.DataFrame(raw_data)
        df.to_csv(save_file_name, mode='a', index=False, header=False,encoding='utf-8-Sig')
    else:
        raw_data = ["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"]
        df = pd.DataFrame(raw_data)
        df.to_csv(save_file_name, mode='a', index=False, header=False,encoding='utf-8-Sig')
    if os.path.exists("TimeoutURL.csv"):#刪除舊資料
        os.remove("TimeoutURL.csv")
if __name__ == '__main__':
    import time
    import os
    import pandas as pd
    import random
    Href_list = Input('./網址/旅遊戶外key.xlsx')#匯入檔案名稱
    save_file_name="旅遊戶外key.csv"#設定存檔名稱
    Create_and_delete_file(save_file_name)
    start = time.time() # 開始測量執行時間
    collector = IPCollector()
    filtered_ips = collector.collect_and_check_ip()
    check=0
    Set_Number=20#設定一次執行數量
    n=0
    m=Set_Number+n
    Index=CcIndex()
    while m<101:
        if check>5:
            #每執行20次重抓Ip
            if check%20==0:
                collector = IPCollector()
                filtered_ips = collector.collect_and_check_ip()
            #每執行10次休息10秒
            if check%10==0:
                time.sleep(10)
        link = Href_list[n:m]
        proxy_ip=[]
        for i in range(len(link)):
            proxy_ip.append(random.choice(filtered_ips))
        MultiProcess1(link,proxy_ip)
        time.sleep(5)
        print("執行",m,"筆資料")
        n=m
        m=m+Set_Number
        check=check+1
        continue
    end = time.time() # 結束測量執行時間
    print("平行後執行時間為 %f 秒" % (end - start))
