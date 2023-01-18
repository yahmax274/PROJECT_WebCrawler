def MultiProcess(n,m,Ip):
    import concurrent.futures
    with concurrent.futures.ProcessPoolExecutor() as executor:
        for result  in executor.map(Main ,Href_list,Ip, chunksize=500):
            try:
                total.append(result )
            except:
                continue 
    return total
def MultiProcess1(link,proxy_ip):  
    with concurrent.futures.ProcessPoolExecutor() as executor:
        ## 更改 main、Main 可改變執行的function
        a = executor.map(main,link,proxy_ip) 
        executor.shutdown(wait=False)
        try:
            for b in a:
                if "https" in b:
                    error.append(b)
                else:   
                    total.append(b)
        except:
            print("null")
    return total
##----------####-----RandomIP-----####----------##
def IpCollect():#收集IP步驟1
    import concurrent.futures
    import requests
    import re
    proxy_ips=[]
    response = requests.get("https://www.sslproxies.org/")
    proxy_ips = re.findall('\d+\.\d+\.\d+\.\d+:\d+', response.text) 
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = [executor.submit(IpCheck, link) for link in proxy_ips]
        for future in concurrent.futures.as_completed(futures):
            try:
                IpUseable.append(future.result())
            except:
                continue 
    print("收集到的IP數量:",len(IpUseable))
    IpGet=IpReCollect()
    return IpGet
def IpCheck(ip):#收集IP步驟2
    import requests
    # Useable=""
    try:
        result = requests.get('https://ip.seeip.org/jsonip?',
			       proxies={'http': ip, 'https': ip},
			       timeout=5)
        print(result.json())
        Useable=ip
    except:
        pass
    return Useable
def IpReCollect():#收集IP步驟3
    import concurrent.futures
    import requests
    import re
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = [executor.submit(IpReCheck, link) for link in IpUseable]
        for future in concurrent.futures.as_completed(futures):
            try:
                IpReUseable.append(future.result())
            except:
                continue 
    print("確定能使用的IP數量:",len(IpReUseable))
    print(IpReUseable)
    return IpReUseable
def IpReCheck(ip):#收集IP步驟4
    import requests
    from bs4 import BeautifulSoup
    from fake_useragent import UserAgent
    try:
        user_agent = UserAgent()
        headers={ 'user-agent': user_agent.random }
        result = requests.get('https://www.momoshop.com.tw/main/Main.jsp',
			       proxies={'http': ip, 'https': ip},
			       timeout=5,
                   headers=headers)
        if result.status_code==200:
            Useable=ip
    except:
        pass
    return Useable
##----------####-----RandomIP-----####----------##
def Input(file):#匯入Excel資料
    from openpyxl import load_workbook
    wb = load_workbook(file)
    # Work Sheet
    ws = wb.get_sheet_by_name('sheet1')
    # Column
    column = ws['A']
    link_list = [column[x].value for x in range(len(column))]
    return (link_list)   
def Output(total):#匯出csv檔
    import pandas as pd
    raw_data ={"app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act": total}
    df = pd.DataFrame(raw_data,columns=["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"])
    df.to_csv("3Ckey.csv",encoding='utf-8-Sig',index=False)
def Main(link,proxy_ip):#主函式
    import requests
    from bs4 import BeautifulSoup
    import datetime
    from fake_useragent import UserAgent
    import random
    import time
    delay_choices = [1,2,3,4,5]  #延遲的秒數
    delay = random.choice(delay_choices)
    time.sleep(delay)
    a=0
    url = link
    ip=proxy_ip
    user_agent = UserAgent()
    headers={ 'user-agent': user_agent.random }
    re = requests.get(url,headers=headers, timeout=10,proxies={'http': f'{ip}', 'https': f'{ip}'})
    soup = BeautifulSoup(re.text,"html.parser")
    #type
    types = soup.find_all("div",{"id":"bt_2_layout_NAV"})
    try:
        for type in types:
            type = type.text
            type = type.replace(" ","")
            type = type.split("\n")
            type = type[2:]
            if type == []:
                continue
            else:
                a=1
                for i in range(4):  
                    if type[i] == "":
                        type[i] ="null"
                type = type[0]+"|"+type[1]+"|"+type[2]+"|"+type[3]+"|"+type[4]
                type_list=type
                
        #brand
        brand = soup.find("a",{"class":"webBrandLink"}).get_text()
        brand_list=brand
        #item_name
        name = soup.find("span",{"id":"osmGoodsName"}).get_text()
        name_list=name
        #prices
        prices = soup.find_all("ul",class_ = "prdPrice")
        if prices ==[]:
                price = "null|null|null"
                price_list=price
        else:
            for price in prices:
                price = price.text
                price = price.replace(" ","")
                price = price.split("\n")
                price = [i.strip() for i in price if i.strip()!=""]
                price_1 =""
                price_2 =""
                price_3 =""
                for i in range(len(price)):
                    if price_1 =="null" or price_1 =="":
                        if "市售價" in price[i]:
                            price_1 = price[i]
                            price_1 = price_1[3:]
                            price_1 = price_1.replace("元","")   
                            price_1 = price_1.replace(",","")
                        else:
                            price_1 = "null"
                    if price_2 =="null" or price_2 =="":
                        if "促銷價" in price[i]:
                            price_2 = price[i]
                            price_2 = price_2[3:]
                            price_2 = price_2.replace("元","")
                            price_2 = price_2.replace("賣貴通報","")
                            price_2 = price_2.replace("下單再折","")
                            price_2 = price_2.replace(",","")
                        else:
                            price_2 = "null"  
                    if price_3 =="null" or price_3 =="":
                        if "折扣後價格" in price[i]:
                            price_3 = price[i]
                            price_3 = price_3[5:]
                            price_3 = price_3.replace("元賣貴通報","")
                            price_3 = price_3.replace(",","") 
                        else:
                            price_3 = "null"  
                price = price_1+"|"+price_2+"|"+price_3
                price_list=price
        #date
        date = datetime.date.today()
        date = str(date)
        date = date.replace("-","")
        date_list=date
        #spec
        spec = soup.find("div",class_ = "vendordetailview specification").get_text()
        spec = spec.replace("\n","")
        spec = spec.replace(" ","")
        spec = spec.strip("TOP")
        spec_list=spec
        #act
        try:
            act = soup.find("ul",class_ = "ineventArea").text
            act = act.replace(" ","")
            act = act.replace("\n","")
            act = act.replace("\xa0","")
            act = act.split("(說明)")
            act = act[:2]
            if act[1] =="":
                act = act[0]+"|null"
            else:
                act = act[0] +"|"+ act[1]       
        except:
            act = "null|null"
        act_list=act
    except:
        index="NoResponse"
        pass
    if a==1:
        index = type_list+"|"+brand_list+"|"+name_list+"|"+price_list+"|"+date_list+"|"+spec_list+"|"+act_list
    else:
        index="null"
    return index
def main(link,proxy_ip):
    import requests
    from bs4 import BeautifulSoup
    import datetime
    from fake_useragent import UserAgent
    import random
    import time
    all = ""
    type_list = []
    brand_list = []
    name_list = []
    price_list = []
    date_list = []
    spec_list = []
    act_list = []
    delay_choices = [1,2,3,4,5]  #延遲的秒數
    delay = random.choice(delay_choices)
    time.sleep(delay)
    url = link
    ip=proxy_ip
    user_agent = UserAgent()
    headers={ 'user-agent': user_agent.random }
    re = requests.get(url,headers=headers, timeout=20,proxies={'http': f'{ip}', 'https': f'{ip}'})
    soup = BeautifulSoup(re.text,"html.parser")
    #type
    types = soup.find_all("div",{"id":"bt_2_layout_NAV"})
    if types != []:
        for type in types:
            type = type.text
            type = type.replace(" ","")
            type = type.split("\n")
            type = type[2:] 
            for i in range(4):  
                if type[i] == "":
                    type[i] ="null"
            type = type[0]+"|"+type[1]+"|"+type[2]+"|"+type[3]+"|"+type[4] 
            type_list.append(type) 
            
    else:
        all = url
    try:    
        #brand
        brand = soup.find("a",{"class":"webBrandLink"}).get_text()
        brand_list.append(brand)
        #item_name
        name = soup.find("span",{"id":"osmGoodsName"}).get_text()
        name_list.append(name)
        #prices
        prices = soup.find_all("ul",class_ = "prdPrice")
        if prices ==[]:
                price = "null|null|null"
                price_list.append(price)
        else:
            for price in prices:
                price = price.text
                price = price.replace(" ","")
                price = price.split("\n")
                price = [i.strip() for i in price if i.strip()!=""]
                price_1 =""
                price_2 =""
                price_3 =""
                for i in range(len(price)):
                    if price_1 =="null" or price_1 =="":
                        if "市售價" in price[i]:
                            price_1 = price[i]
                            price_1 = price_1[3:]
                            price_1 = price_1.replace("元","")   
                            price_1 = price_1.replace(",","")
                        else:
                            price_1 = "null"
                    if price_2 =="null" or price_2 =="":
                        if "促銷價" in price[i]:
                            price_2 = price[i]
                            price_2 = price_2[3:]
                            price_2 = price_2.replace("元","")
                            price_2 = price_2.replace("賣貴通報","")
                            price_2 = price_2.replace("下單再折","")
                            price_2 = price_2.replace(",","")
                        else:
                            price_2 = "null"  
                    if price_3 =="null" or price_3 =="":
                        if "折扣後價格" in price[i]:
                            price_3 = price[i]
                            price_3 = price_3[5:]
                            price_3 = price_3.replace("元賣貴通報","")
                            price_3 = price_3.replace(",","") 
                        else:
                            price_3 = "null"  
                price = price_1+"|"+price_2+"|"+price_3
                price_list.append(price)
        #date
        date = datetime.date.today()
        date = str(date)
        date = date.replace("-","")
        date_list.append(date)
        #spec
        spec = soup.find("div",class_ = "vendordetailview specification").get_text()
        spec = spec.replace("\n","")
        spec = spec.replace(" ","")
        spec = spec.strip("TOP")
        spec_list.append(spec)
        #act
        try:
            act = soup.find("ul",class_ = "ineventArea").text
            act = act.replace(" ","")
            act = act.replace("\n","")
            act = act.replace("\xa0","")
            act = act.split("(說明)")
            act = act[:2]
            if act[1] =="":
                act = act[0]+"|null"
            else:
                act = act[0] +"|"+ act[1]       
        except:
            act = "null|null"
        act_list.append(act)
        for i in range(len(type_list)):  
            all = type_list[i]+"|"+brand_list[i]+"|"+name_list[i]+"|"+price_list[i]+"|"+date_list[i]+"|"+spec_list[i]+"|"+act_list[i]
        return(all)
    except:    
        return(all)
def CcIndex():
    Index=len(Href_list)+300
    return Index
if __name__ == '__main__':
    import time
    import concurrent.futures
    import random
    all_list = []
    Href_list=[]
    total=[]
    error=[]
    Check_list=[]
    proxy_ips=[]
    valid_ips=[]
    IpUseable=[]
    IpReUseable=[]
    Index=[]
    start = time.time() # 開始測量執行時間
    thiscycle=time.time()# 開始測量本迴圈執行時間(僅用一次)
    start_time = time.ctime(start)
    print("開始執行時間：", start_time)
    Href_list = Input('3Ckey.xlsx')#匯入檔案名稱
    Index=CcIndex()
    Ip=IpCollect()
    print("IP_Len:",len(Ip))
    check=0
    Set_Number=10#一次執行數量
    n=0
    m=Set_Number+n
    while m<Index:
        if check==1:##debug用
            if len(total)<5:
                print("中止執行!!!!!!!")
                break
        if len(total)>11:##debug用
            break
        if check>5:
            #爬蟲被發現時結束程式
            if ((Check_list[check-1]+Check_list[check-2]+Check_list[check-3])/3)==Check_list[check-1]:
                print("終止執行")
                break
            #爬蟲精確度過低時重新抓取IP
            if Check_list[check-1]!=Check_list[check-2]:
                if Check_list[check-2]!=Check_list[check-3]:
                    if Check_list[check-3]!=Check_list[check-4]:
                        if (Check_list[check-1]-Check_list[check-2])<Set_Number*0.4:
                            print("重抓Ip", check)
                            Ip.clear()
                            Ip=IpCollect()
                            print("IP_Len:",len(Ip))
            #每執行10次暫時存檔
            if check%10==0:
                now = time.time()
                print("執行10次Ip重抓,執行第", check)
                Ip.clear()
                Ip=IpCollect()
                print("IP_Len:",len(Ip))
                print("執行第",check*Set_Number,"次時間為 %f 秒" % (now - start))
                print("暫存一次")
                Output(total)
        link = Href_list[n:m] 
        proxy_ip=[]
        for i in range(len(link)):
            proxy_ip.append(random.choice(Ip))
        MultiProcess1(link,proxy_ip)
        time.sleep(5)
        n=m
        m=m+Set_Number
        print(len(total))
        check=check+1
        Check_list.append(len(total))
        continue
    print("遺失數：",(check*Set_Number)-len(total))
    Output(total)
    end = time.time() # 結束測量執行時間
    end_time = time.ctime(end)
    print("執行數量",check*Set_Number)
    print("執行結束時間：", end_time)
    print("執行時間為 %f 秒" % (end - start))
