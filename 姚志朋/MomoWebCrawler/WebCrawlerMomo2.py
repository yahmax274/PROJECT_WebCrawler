import requests
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
import time
from webdriver_manager.chrome import ChromeDriverManager
##-----------##新增編輯CSV
wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
s1 = wb["Sheet"]        # 取得工作表名稱為「工作表1」的內容
Tittle=["商品名稱","原價","折扣","特價","剩餘數量","時間"]
for i in range(len(Tittle)):
    y=i+1
    c = s1 .cell(row=1, column=y)
    c.value =Tittle[i]
##-----------##網址抓取
url=("https://www.momoshop.com.tw/ajax/promotionEvent_CustExclbuy.jsp")
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)Chrome/106.0.0.0 Safari/537.36'}
my_data = {'key1': 'value1', 'key2': 'value2',"showType":1}
r = requests.post(url,headers=headers, data = my_data)
print(r.status_code)
# print(r.text)
##------##
content = r.text
soup = BeautifulSoup(content, "html.parser")
# ##-----------##商品名稱
# ProductBrand=[]
# ProductName=[]
# for link in soup.find_all("div", class_= "brand",id="gdsBrand_1"):
#     ProductBrand.append(link.get_text("id"))
# for link in soup.find_all("div", class_= "brand2",id="gdsName_1"):
#     ProductName.append(link.get_text("id"))
# for i in range(len(ProductBrand)):
#     y=i+2
#     c = s1 .cell(row=y, column=1)
#     c.value =(ProductBrand[i])+(ProductName[i])
# ##-----------##原價
# OriginalPrice =[]
# for link in soup.find_all("span", id="sPrice_1"):
#     OriginalPrice.append(link.get_text("id"))
# for i in range(len(OriginalPrice)):
#     y=i+2
#     c = s1 .cell(row=y, column=2)
#     c.value =OriginalPrice[i]
# ##-----------##折扣
# Discount =[]
# for link in soup.find_all("span", id="discAmt_1"):
#     Discount.append(link.get_text("id"))
# for i in range(len(Discount)):
#     y=i+2
#     c = s1 .cell(row=y, column=3)
#     c.value =(Discount[i])+"折"
# ##-----------##特價
# SpecialOffer =[]
# for link in soup.find_all("span" ,id="nPrice_1"):
#     SpecialOffer.append(link.get_text("id"))
# for i in range(len(SpecialOffer)):
#     y=i+2
#     c = s1 .cell(row=y, column=4)
#     c.value =SpecialOffer[i]
# ##-----------##剩餘數量
# TheRemainingAmount =[]
# for link in soup.find_all("span", id="gdsStock_1"):
#     TheRemainingAmount.append(link.get_text("id"))
# for i in range(len(TheRemainingAmount)):
#     y=i+2
#     c = s1 .cell(row=y, column=5)
#     c.value =TheRemainingAmount[i]
# ##-----------##時間
# Time =[]
# for link in soup.find_all("div" ,class_="time"):
#     for i in link.find("span"):
#         time_text = i.text
#         Time.append(time_text)
# y=0
# cc=0
# for link in soup.find_all("ul" ,class_="product_Area"):
#     CcTime=Time[cc]
#     for i in link.find(class_="box1"):
#         y=y+1
#         c = s1 .cell(row=y+1, column=6)
#         c.value =CcTime
#     cc=cc+1
##-----------##抓網址
Href=[]
for link in soup.find_all('a'):
    Href.append(link.get('href').replace("//",""))
# print(Href)
print(Href[0])
url=Href[0]
driver = webdriver.Chrome(ChromeDriverManager().install())
print(url)
driver.get(url)
# driver.get("https://tw.yahoo.com/")
time.sleep(5)
print(driver.title)
# response = webdriver.request('post', url,header=headers, data = my_data)
# print(response)
driver.close() # 關閉瀏覽器視窗


##-----------##儲存CSV
wb.save('Momo2.xlsx')
