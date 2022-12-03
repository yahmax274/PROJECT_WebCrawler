def MultThread(n,m):
    import concurrent.futures
    # with concurrent.futures.ThreadPoolExecutor() as executor:
    with concurrent.futures.ProcessPoolExecutor() as executor:
        executor.map(Main, Href_list[n:m])
def MultiProcess(n,m):
    with concurrent.futures.ProcessPoolExecutor() as executor:
        for result in executor.map(Main, Href_list[n:m], chunksize=50):
            total.append(result)
            try:
                total.append(result)
            except:
                total.append("Error")
                error.append("Error")
                continue 
    return total,error
def MultiProcess1(n,m,Ip):
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = [executor.submit(Main, link,Ip) for link in Href_list[n:m]]
        for future in concurrent.futures.as_completed(futures):
            try:
                total.append(future.result())
            except:
                # total.append("ReadTimeOut")
                error.append("ReadTimeOut")
                continue 
    return total,error
def ChangeIp():
    import requests
    import re
    print("重新抓取Proxy IP")
    response = requests.get("https://www.sslproxies.org/")
    proxy_ips = re.findall('\d+\.\d+\.\d+\.\d+:\d+', response.text)  #「\d+」代表數字一個位數以上
    valid_ips = []
    for index, ip in enumerate(proxy_ips):
        try:
            if index <= 30: 
                result = requests.get('https://ip.seeip.org/jsonip?',
                                    proxies={'http': ip, 'https': ip},
                                    timeout=5)
                print(result.json())
                valid_ips.append(ip)
        except:
            pass
    return valid_ips
def Input(file):
    from openpyxl import load_workbook
    wb = load_workbook(file)
    # Work Sheet
    ws = wb.get_sheet_by_name('sheet1')
    # Column
    column = ws['A']
    link_list = [column[x].value for x in range(len(column))]
    return (link_list)   
#匯出csv檔
def Output(total):    
    import pandas as pd
    raw_data ={"app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act": total}
    df = pd.DataFrame(raw_data,columns=["app|big|mid|small|category|brand|item_name|market_price|sale_price|discount_price|date|item_specification|act"])
    df.to_csv("momo3.csv",encoding='utf-8-Sig',index=False)
    
def Main(link,Ip):
    import requests
    from bs4 import BeautifulSoup
    import datetime
    from fake_useragent import UserAgent
    import random
    import time
    delay_choices = [1,2,3,4,5]  #延遲的秒數
    delay = random.choice(delay_choices)
    time.sleep(delay)
    a=0
    url = link
    try:
        proxy_ip = random.choice(Ip)  # 隨機取得Proxy IP
        # print(f'使用的Proxy IP：{proxy_ip}')
    except:
        pass
    user_agent = UserAgent()
    headers={ 'user-agent': user_agent.random }
    re = requests.get(url,headers=headers, timeout=10,proxies={'http': f'{proxy_ip}', 'https': f'{proxy_ip}'})
    soup = BeautifulSoup(re.text,"html.parser")
    #type
    types = soup.find_all("div",{"id":"bt_2_layout_NAV"})
    try:
        for type in types:
            type = type.text
            type = type.replace(" ","")
            type = type.split("\n")
            type = type[2:]
            if type == []:
                continue
            else:
                a=1
                for i in range(4):  
                    if type[i] == "":
                        type[i] ="null"
                type = type[0]+"|"+type[1]+"|"+type[2]+"|"+type[3]+"|"+type[4]
                type_list=type
                
        #brand
        brand = soup.find("a",{"class":"webBrandLink"}).get_text()
        brand_list=brand
        #item_name
        name = soup.find("span",{"id":"osmGoodsName"}).get_text()
        name_list=name
        #prices
        prices = soup.find_all("ul",class_ = "prdPrice")
        if prices ==[]:
                price = "null|null|null"
                price_list=price
        else:
            for price in prices:
                price = price.text
                price = price.replace(" ","")
                price = price.split("\n")
                price = [i.strip() for i in price if i.strip()!=""]
                price_1 =""
                price_2 =""
                price_3 =""
                for i in range(len(price)):
                    if price_1 =="null" or price_1 =="":
                        if "市售價" in price[i]:
                            price_1 = price[i]
                            price_1 = price_1[3:]
                            price_1 = price_1.replace("元","")   
                            price_1 = price_1.replace(",","")
                        else:
                            price_1 = "null"
                    if price_2 =="null" or price_2 =="":
                        if "促銷價" in price[i]:
                            price_2 = price[i]
                            price_2 = price_2[3:]
                            price_2 = price_2.replace("元","")
                            price_2 = price_2.replace("賣貴通報","")
                            price_2 = price_2.replace("下單再折","")
                            price_2 = price_2.replace(",","")
                        else:
                            price_2 = "null"  
                    if price_3 =="null" or price_3 =="":
                        if "折扣後價格" in price[i]:
                            price_3 = price[i]
                            price_3 = price_3[5:]
                            price_3 = price_3.replace("元賣貴通報","")
                            price_3 = price_3.replace(",","") 
                        else:
                            price_3 = "null"  
                price = price_1+"|"+price_2+"|"+price_3
                price_list=price
        #date
        date = datetime.date.today()
        date = str(date)
        date = date.replace("-","")
        date_list=date
        #spec
        spec = soup.find("div",class_ = "vendordetailview specification").get_text()
        spec = spec.replace("\n","")
        spec = spec.replace(" ","")
        spec = spec.strip("TOP")
        spec_list=spec
        #act
        try:
            act = soup.find("ul",class_ = "ineventArea").text
            act = act.replace(" ","")
            act = act.replace("\n","")
            act = act.replace("\xa0","")
            act = act.split("(說明)")
            act = act[:2]
            if act[1] =="":
                act = act[0]+"|null"
            else:
                act = act[0] +"|"+ act[1]       
        except:
            act = "null|null"
        act_list=act
    except:
        index="NoResponse"
        pass
    if a==1:
        index = type_list+"|"+brand_list+"|"+name_list+"|"+price_list+"|"+date_list+"|"+spec_list+"|"+act_list
    else:
        index="null"
    return index
def CcIndex():
    Index=len(Href_list)+50
    return Index
if __name__ == '__main__':
    import time
    import concurrent.futures
    import random
    all_list = []
    Href_list=[]
    total=[]
    error=[]
    Check_list=[]
    start = time.time() # 開始測量執行時間
    start_time = time.ctime(start)
    print("開始執行時間：", start_time)
    Href_list = Input('3Ckey.xlsx')#匯入檔案名稱
    Index=CcIndex()
    Ip=ChangeIp()
    check=0
    Set_Number=50#一次執行數量
    n=0
    m=Set_Number+n
    while m<Index:
        if check>5:
            if ((Check_list[check-1]+Check_list[check-2]+Check_list[check-3])/3)==Check_list[check-1]:
                break
            if check%10==0:
#                 delay_choices = [30,45,60]  #延遲的秒數
#                 delay = random.choice(delay_choices)
                print("執行10次休息：Ip重抓,執行第", check)
#                 time.sleep(delay)
                Ip.clear()
                Ip=ChangeIp()
                print("IP_Len:",len(Ip))
        # MultiProcess(n,m)
        MultiProcess1(n,m,Ip)
        time.sleep(5)
        n=m
        m=m+Set_Number
        print(len(total))
        check=check+1
        Check_list.append(len(total))
        continue
    print("實際得到：",len(total)-len(error))
    Output(total)
    end = time.time() # 結束測量執行時間
    end_time = time.ctime(end)
    print("執行結束時間：", end_time)
    print("執行時間為 %f 秒" % (end - start))
